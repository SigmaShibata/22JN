import openpyxl as px

wb = px.load_workbook("xlsx/seiseki_fix.xlsx")
ws = wb.active
max_row = ws.max_row
max_column = ws.max_column

subjects = []
for i in range(2, ws.max_column + 1):
    subjects.append(ws.cell(row=1, column=i).value)

subject = input("科目を入力 --> ")
is_found = False
while not is_found:
    for i in range(2, ws.max_column + 1):
        if subject == ws.cell(row=1, column=i).value:
            is_found = True
            wb.create_sheet(subject)
            ws_new = wb[subject]
            judge = int(input("合格点を入力-->"))
            for j in range(2, ws.max_row + 1):
                ws_new.cell(row=j-1, column=1).value = ws.cell(row=j, column=1).value
                if ws.cell(row=j, column=i).value >= judge:
                    ws_new.cell(row=j-1, column=2).value = "合格"
                else:
                    ws_new.cell(row=j-1, column=2).value = "不合格"
            break
    if not is_found:
        print("入力された科目は登録されていません。")
        subject = input("科目を入力 --> ")

wb.save("xlsx/seiseki_fix.xlsx")
wb.close()