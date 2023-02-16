import openpyxl as px
menu = 0
while menu != 9:
    wb = px.load_workbook("xlsx/shohin.xlsx")
    ws = wb.active

    codes = []
    goods = []
    for i in range(2, ws.max_row):
        code = ws.cell(row=i, column=1).value
        codes.append(code)
        name = ws.cell(row=i, column=2).value
        goods.append(name)
    
    menu = int(input("一覧:1 追加:2 削除:3 終了:9 --> "))
    if menu == 1:
        for i in range(1, ws.max_row+1):
            print(f"{ws.cell(row=i, column=1).value}\t{ws.cell(row=i, column=2).value}")
    if menu == 2:
        code = int(input("商品コード --> "))
        if code in codes:
            print(f"商品コード{code}: {goods[codes.index(code)]}は既に登録されています。")
        else:
            name = input("商品名 --> ")
            ws.cell(row=ws.max_row+1, column=1).value = code
            ws.cell(row=ws.max_row, column=2).value = name
            print(f"商品コード{code}: {name}を登録しました。")
            
    if menu == 3:
        code = int(input("商品コード --> "))
        if code in codes:
            num = codes.index(code)
            for i in range(num+2, ws.max_row):
                ws.cell(row=i, column=1).value = ws.cell(row=i+1, column=1).value
                ws.cell(row=i, column=2).value = ws.cell(row=i+1, column=2).value
            ws.cell(row=ws.max_row, column=1).value = None
            ws.cell(row=ws.max_row, column=2).value = None
            print(f"商品コード{code}: {goods[codes.index(code)]}を削除しました。")
        else:
            print(f"商品コード{code} は登録されていません。")
    wb.save("xlsx/shohin.xlsx")
    wb.close()

print("終了しました。")